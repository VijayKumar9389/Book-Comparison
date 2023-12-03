from pickle import FALSE
from openpyxl import *

def extractBook(path):

    content = []
    book = load_workbook(path)
    sheet = book.active
    row_count = sheet.max_row
    column_count = sheet.max_column

    for i in range(1, row_count + 1):
        row = []
        for j in range(1, column_count + 1):
            if(sheet.cell(row=i, column=j).value == None):
                row.append('')
            else:
                row.append(sheet.cell(row=i, column=j).value)
        content.append(row)
    return content

def compareRows(rowOne, rowTwo):

    for i in range(1, len(rowOne)):
        if (rowOne[i] == rowTwo[i]):
            print('')
        else:
            valueOne = rowOne[i].split()
            valueTwo = rowTwo[i].split()
            listting = set(valueOne + valueTwo)
            compareValues(valueOne, valueTwo)

def compareValues(valueOne, valueTwo):

    keep = ['--keep--']
    add = ['--add--']
    delete = ['--delete--']

#Checks to see what characters are the same or new compared to the old book
    for y in range(0, len(valueTwo)):

        stringTwo = valueTwo[y]

        #checks if character is in the old book
        if stringTwo in valueOne: 
            keep.append(stringTwo)
        #checks if character is in the new book
        if stringTwo not in valueOne: 
            add.append(stringTwo)

#checks if characters where removed in the new book
    for i in range(0, len(valueOne)):
        
        stringOne = valueOne[i]

        if stringOne not in (valueTwo):
            delete.append(stringOne)


    testing = [*keep, *add, *delete]
    print(testing)

def Compare(itemOne, itemTwo):

    if str(itemOne) == str(itemTwo):
        return True
    return False